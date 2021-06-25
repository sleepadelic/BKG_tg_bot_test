import datetime
import openpyxl
import settings
from PIL import Image

import issue_combiner
import Models
import secret
from openpyxl import Workbook
from dadata import Dadata


def excel_export_main():
    issues = combine_reports()
    print("combined")
    issues = select_issues_by_date(datetime.datetime.now().date(), issues)
    print("sorted")
    issues = load_addresses(issues)
    print("Addresses loaded")
    issues = img_relative_path(issues)
    print("Img paths fixed")
    saved_yaml_file(issues)
    print("Saved .yaml")
    export_to_xlsx(issues)
    print("Report saved")


def export_to_xlsx(issues, filename=f'../{settings.report_files_directory}{str(datetime.datetime.now().date())}.xlsx'):
    """
    Создаёт отчёт из подготовленного yaml файла
    :param filename: Путь для сохранения отчёта
    """
    wb = Workbook()
    ws = wb.active
    Create_headlines(ws)
    # change column size for image
    ws.column_dimensions['G'].width = 450
    row_position = 2
    iss: Models.Issue
    for iss in issues:
        ws[f'A{str(row_position)}'] = str(row_position - 1)
        ws[f'B{str(row_position)}'] = str(iss.send_time.date())
        ws[f'C{str(row_position)}'] = iss.type
        ws[f'D{str(row_position)}'] = iss.address
        ws[f'E{str(row_position)}'] = iss.description
        ws[f'F{str(row_position)}'] = iss.geo

        img = openpyxl.drawing.image.Image('../' + iss.image)
        img.anchor = f'G{str(row_position)}'
        img.width = img.width / 3
        img.height = img.height / 3
        ws.add_image(img)
        ws.row_dimensions[row_position].height = 240

        row_position += 1
    wb.save(filename=filename)


def Create_headlines(ws):
    ws['A1'] = "№"
    ws['B1'] = "Дата"
    ws['C1'] = "Тип обращения"
    ws['D1'] = "Адрес"
    ws['E1'] = "Описание"
    ws['F1'] = "Геопозиция"
    ws['G1'] = "Фото"


def combine_reports(folder_path="../data/"):
    """
    Объединяет отчёты в один файл
    """
    issues = issue_combiner.open_and_load_to_array(folder_path)
    return issues


def select_issues_by_date(date, issues):
    """
    :param date: дата в формате год-месяц-день
    :param issues: список issues
    :return: выбранные issues по дате
    """
    selected_issues = []
    iss: Models.Issue
    for iss in issues:
        if date == iss.send_time.date():
            selected_issues.append(iss)

    return selected_issues


def select_issues_by_type(type: str, issues):
    """
    :param type: выбор с клавиатуры, строка
    :param issues: список issues
    :return: выбранные issues по типу
    """
    selected_issues = []
    iss: Models.Issue
    for iss in issues:
        if type == iss.type:
            selected_issues.append(iss)

    return selected_issues


def select_issues_by_type_and_date(date: str, type: str, issues):
    """

    :param date: дата в формате год-месяц-день
    :param type: выбор с клавиатуры, строка
    :param issues: список issues
    :return: выбранные issues по дате и типу
    """
    selected_issues = select_issues_by_date(date, issues)
    selected_issues = select_issues_by_type(type, selected_issues)
    return selected_issues


def saved_yaml_file(issues):
    issue_combiner.save_to_yml(issues, f"../{settings.report_files_directory}combined_export.yaml")


def img_relative_path(issues):
    """
    Делает путь к изображениям относительным
    """
    iss: Models.Issue
    for iss in issues:
        iss.image = iss.image.replace('/home/danil0111/bkg_bot/', '')
    return issues


def load_addresses(issues):
    """
    Загружает адреса по гео-координатам
    :return:
    """
    iss: Models.Issue
    for iss in issues:
        try:
            if (iss.geo != None):
                geo = iss.geo.split(',')
                iss.address = reverse_geocode(geo[0], geo[1])[0]['unrestricted_value']
        except Exception as exc:
            print(f"No address in {iss.geo} {exc}")
    return issues


def reverse_geocode(lat, lon):
    """
    Преобразует координаты в адрес
    :rtype: str address
    """
    with Dadata(secret.dadata_token, secret.dadata_secret) as dadata:
        return dadata.geolocate(name='address', lat=lat, lon=lon)


if __name__ == '__main__':
    excel_export_main()
